#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Oct 31 14:12:06 2022

@author: user
"""

from pann.models import Cnn14_DecisionLevelMaxMels, Cnn14_DecisionLevelMax, ResNet38Mels, ResNet38, Cnn14Mels, Cnn14
import torch
import numpy as np
import openpyxl
from pathlib import Path
from sklearn import preprocessing
import torch.nn.functional as F
from pathlib import Path
import pandas as pd
import utils.util as ut
import utils.bands_transform as bt

class PannMelInference():
    def __init__(self, device=torch.device("cpu"), verbose=False, pann_type='ResNet38'):
        self.name = "PANN"
        self.sample_rate = 32000
        self.window_size = 1024
        self.hop_size = 320
        self.mel_bins = 64
        self.fmin = 50
        self.fmax = 14000
        # Load labels from the first column of the Excel file using pandas
        df = pd.read_excel('./utils/audioset_tvb.xlsx', usecols=[0])
        self.labels_str = df.iloc[:, 0].dropna().tolist()
        self.le = preprocessing.LabelEncoder()
        self.labels_enc = self.le.fit_transform(self.labels_str)
        self.labels_enc = torch.from_numpy(self.labels_enc)

        self.n_labels = len(self.labels_str)
        self.device = device
        self.type = pann_type
        self.n_embedding = 2048

        if pann_type == 'CNN14LevelMax':
            #model that takes Mel spectrogram as input
            self.model = Cnn14_DecisionLevelMaxMels(sample_rate=self.sample_rate, window_size=self.window_size, 
                hop_size=self.hop_size, mel_bins=self.mel_bins, fmin=self.fmin, fmax=self.fmax, 
                classes_num=self.n_labels)
            #model that takes audio as input
            self.full_model =  Cnn14_DecisionLevelMax(sample_rate=self.sample_rate, window_size=self.window_size, 
                hop_size=self.hop_size, mel_bins=self.mel_bins, fmin=self.fmin, fmax=self.fmax, 
                classes_num=self.n_labels)
            
        if pann_type == 'ResNet38':
            #model that takes Mel spectrogram as input
            self.model = ResNet38Mels(sample_rate=self.sample_rate, window_size=self.window_size, 
                hop_size=self.hop_size, mel_bins=self.mel_bins, fmin=self.fmin, fmax=self.fmax, 
                classes_num=self.n_labels)
            
            #model that takes audio as input
            self.full_model =  ResNet38(sample_rate=self.sample_rate, window_size=self.window_size, 
                hop_size=self.hop_size, mel_bins=self.mel_bins, fmin=self.fmin, fmax=self.fmax, 
                classes_num=self.n_labels)

        if pann_type == 'CNN14':
            features_list = ["2048", "logits"]

            #model that takes Mel spectrogram as input
            self.model = Cnn14Mels(features_list=features_list, sample_rate=self.sample_rate, window_size=self.window_size, 
                hop_size=self.hop_size, mel_bins=self.mel_bins, fmin=self.fmin, fmax=self.fmax, 
                classes_num=self.n_labels)
            
            #model that takes audio as input
            self.full_model =  Cnn14(features_list=features_list, sample_rate=self.sample_rate, window_size=self.window_size, 
                hop_size=self.hop_size, mel_bins=self.mel_bins, fmin=self.fmin, fmax=self.fmax, 
                classes_num=self.n_labels)
            
        ###############
        #models loading
        if pann_type == 'CNN14LevelMax':
            self.checkpoint_path = Path().absolute() / 'pann' / 'ckpt' / 'Cnn14_DecisionLevelMax_mAP=0.385.pth'
        if pann_type == 'ResNet38':
            self.checkpoint_path = Path().absolute() / 'pann' / 'ckpt' / 'ResNet38_mAP=0.434.pth'
        if pann_type == 'CNN14':
            self.checkpoint_path = Path().absolute() / 'pann' / 'ckpt' / 'Cnn14_mAP=0.431.pth'

        checkpoint = torch.load(self.checkpoint_path, map_location=device)
        self.full_model.load_state_dict(checkpoint['model'])
        
        full_model_dict = self.full_model.state_dict()
        model_dict = self.model.state_dict()
        
        # 1. filter out unnecessary keys
        full_model_dict = {k: v for k, v in full_model_dict.items() if k in model_dict}
        # 2. overwrite entries in the existing state dict
        model_dict.update(full_model_dict) 
        # 3. load the new state dict
        self.model.load_state_dict(full_model_dict)
        self.model.to(device)

        if verbose:
            print('PANN Parameters')
            ut.count_parameters(self.model)
        
        ###############
        #sub-labels
        sub_classes_path = './utils/audioset_tvb.xlsx'
        self.sub_classes_dict = open_subclasses_dict(sub_classes_path)
        self.labels_tvb_str = [label for label in self.labels_str if self.sub_classes_dict[label] in ['t', 'v', 'b'] ]

        self.labels_tvb_enc = self.le.transform(self.labels_tvb_str)
        self.labels_tvb_enc = torch.from_numpy(self.labels_tvb_enc)
        
        self.n_labels_tvb = len(self.labels_tvb_enc)

        ###############
        #thresholds
        #file = open('./pann/opt_thres.pkl', 'rb')
        #self.threshold = np.array(pickle.load(file))
        self.threshold = [0.2] * self.n_labels

        self.mels_tr = bt.PANNMelsTransform(flen_tho=4096, device=device)

    def simple_inference(self, x, filter_classes=True, softmax=False, no_grad=True, mean=True):
        # Forward
        if no_grad:
            with torch.no_grad():
                self.model.eval()
                batch_output_dict = self.model(x, None)
        else:
            #self.model.eval()
            #with torch.no_grad():
            self.model.eval()
            batch_output_dict = self.model(x, None)
        
        if mean:
            #logits = batch_output_dict['framewise_output'].mean(dim=1)
            logits = batch_output_dict['clipwise_output']
            # logits = torch.unsqueeze(logits, dim=1)
        else:
            logits = batch_output_dict['framewise_output']
        
        return (logits)

    def get_embedding(self, x, filter_classes=True, softmax=False, no_grad=True):
        # Forward
        if no_grad:
            with torch.no_grad():
                self.model.eval()
                output_dict = self.model(x, None)
        else:
            #self.model.eval()
            #with torch.no_grad():
            self.model.eval()
            output_dict = self.model(x, None)
        
        if 'CNN14' in self.type:
            emb = output_dict["2048"]
        else:
            emb = output_dict["embedding"]
        
        return (emb)
    
    def inference(self, x, filter_classes=True, softmax=False, no_grad=True):
        # Forward
        if no_grad:
            with torch.no_grad():
                self.model.eval()
                batch_output_dict = self.model(x, None)
        else:
            #self.model.eval()
            #with torch.no_grad():
            self.model.eval()
            batch_output_dict = self.model(x, None)
        
        #take the mean prediction of every time frame in the spectrogram
        #logits = batch_output_dict['framewise_output'].mean(dim=1)
        logits = batch_output_dict['clipwise_output']

        #framewise_output = batch_output_dict['framewise_output'].data[0]
        """(time_steps, classes_num)"""
        
        #print('Sound event detection result (time_steps x classes_num): {}'.format(
        #    framewise_output.shape))
        
        logits_tvb = torch.Tensor([])
        
        labels_enc = self.labels_enc
        if filter_classes == True:
            labels_enc = self.labels_tvb_enc
            labels_tvb_enc_indices = [k for k in range(len(self.labels_enc)) if self.sub_classes_dict[self.labels_str[k]] in ['t', 'v', 'b'] ]
            logits_tvb = logits[:, labels_tvb_enc_indices]
        
        if softmax:
            logits = F.log_softmax(logits, dim=1)
            logits_tvb = F.log_softmax(logits_tvb, dim=1)

        return(logits, logits_tvb) 

    def logit_to_labels(self, input, tvb=False):
        
        #average over a whole file
        logits_tvb = input.mean(dim=1)
        
        #use for showing top_results
        if tvb:
            labels_enc = self.labels_tvb_enc
        else:
            labels_enc = self.labels_enc

        #labels_enc = labels_enc.to(self.device)

        sorted_indexes = torch.flip(torch.argsort(logits_tvb), dims=[1])

        top_k = 1  # Show top results
 
        labels_enc_top = labels_enc[sorted_indexes[:, 0 : top_k]]
        top_result_mat = logits_tvb[:, sorted_indexes[:, 0 : top_k]]    
        # """(time_steps, top_k)"""

        labels_enc_top = labels_enc_top

        labels_enc_top = labels_enc_top.flatten()
        labels_str_top = self.le.inverse_transform(labels_enc_top)

        return(labels_str_top)

    def batch_logit_to_tvb(self, input, top_k=10):
        """
        expects an input of (n_frames, labels) of numpy array
        """

        #with numpy
        sorted_indices = np.argsort(input, axis=1)[:, ::-1]
        #with torch
        # sorted_indexes = torch.flip(torch.argsort(logits_tvb), dims=[1])

        top_indices = sorted_indices[ :, 0 : top_k]

        #307:car, 300: traffic, 0: speech, 111: bird
        #with numpy
        t = np.expand_dims((top_indices == 307).any(axis=1), axis=1)
        v = np.expand_dims((top_indices == 0).any(axis=1), axis=1)
        b = np.expand_dims((top_indices == 111).any(axis=1), axis=1)

        #with numpy
        tvb_predictions = np.concatenate((t, v, b), axis=1)
        #with torch
        # contains_values = torch.cat((t_label, v_label, b_label), dim=1).float()

        #with numpy
        tvb_predictions_avg = tvb_predictions.mean(axis=0)
        #with torch
        # labels_str_top = contains_values.mean(dim=0)

        #with numpy
        tvb_predictions_avg = np.expand_dims(tvb_predictions_avg, axis=0)        

        return(tvb_predictions_avg)

    def logit_to_logit_tvb(self, logits):
        labels_enc = self.labels_tvb_enc
        labels_tvb_enc_indices = [k for k in range(len(self.labels_enc)) if self.sub_classes_dict[self.labels_str[k]] in ['t', 'v', 'b'] ]
        logits_tvb = logits[:,:, labels_tvb_enc_indices]
        return(logits_tvb)

def open_subclasses_dict(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook.active
    
    first_row = [] # The row where we stock the name of the column
    for col in range(1,2):
        first_row.append( worksheet.cell(1,col).value )
    # tronsform the workbook to a list of dictionnary
    sub_classes_dict = {}
    for row in range(2, worksheet.max_row+1):
        sub_classes_dict[worksheet.cell(row,1).value] = worksheet.cell(row,2).value
    return(sub_classes_dict)